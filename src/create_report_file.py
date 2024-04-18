from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import datetime
import os
import hashlib
from .config import incompatible_pref
from openpyxl.styles import Alignment


class report:
    """tworzenie raportu z importu"""

    def __init__(self):
        #podane tagi sa rozpoznawane jako zgodne z rozporzadzeniem, slownik mowi dodatkowo jaki przedrostek dopisywac gdy wystapi dany tag (np dopisze OT_ do poczatekGorySkarpy)
        self.a = 1

    def controlSum(self, file_path):
        #filename = input(file_path)
        start = datetime.datetime.now()
        sha256_hash = hashlib.sha256()

        with open(file_path, 'rb') as f:
            # Read and update hash string value in blocks of 4K
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        return sha256_hash

    def copyRange(self, startCol, startRow, endCol, endRow, sheet):
        """Funkcja do kopiowania zakresu komorek"""
        styleAll = []
        rangeSelected = []
        allMergedCells = sheet.merged_cells.ranges
        mergedCells = []

        for ranges in allMergedCells:
            if ranges.min_col >= startCol and ranges.min_row >= startRow and ranges.max_col <= endCol and ranges.max_row <= endRow:
                bounds_relative = [ranges.bounds[0] + 1 - startCol, ranges.bounds[1] + 1 - startRow,
                                   ranges.bounds[2] + 1 - startCol, ranges.bounds[3] + 1 - startRow]
                mergedCells.append(bounds_relative)

        # Loops through selected Rows
        for row in range(startRow, endRow + 1, 1):
            # Appends the row to a RowSelected list
            rowSelected = []
            styleLine = []
            for col in range(startCol, endCol + 1, 1):
                styleLine.append([])
                rowSelected.append(sheet.cell(row=row, column=col).value)
                # if sheet.cell(row = row, column = col).has_style:

                # copy style from current cell
                styleLine[-1].append(copy(sheet.cell(row=row, column=col).font))
                styleLine[-1].append(copy(sheet.cell(row=row, column=col).border))
                styleLine[-1].append(copy(sheet.cell(row=row, column=col).fill))
                styleLine[-1].append(copy(sheet.cell(row=row, column=col).number_format))
                styleLine[-1].append(copy(sheet.cell(row=row, column=col).protection))
                styleLine[-1].append(copy(sheet.cell(row=row, column=col).alignment))

            # Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
            styleAll.append(styleLine)
        return rangeSelected, styleAll, mergedCells

    def pasteRange(self, startCol, startRow, endCol, endRow, sheetReceiving, copiedData, style, mergedCells=[],
                   include_value=True):
        """funkcja do wklejania zakresu komorek"""
        # Paste range
        # Paste data from copyRange into template sheet

        if mergedCells:
            for merged in mergedCells:
                boundary_new = get_column_letter(merged[0] + startCol - 1) + str(
                    merged[1] + startRow - 1) + ':' + get_column_letter(merged[2] + startCol - 1) + str(
                    merged[3] + startRow - 1)
                sheetReceiving.merge_cells(boundary_new)
        countRow = 0
        for row in sheetReceiving.iter_rows(min_row=startRow, max_row=endRow, min_col=startCol, max_col=endCol):
            countCol = 0
            for cell in row:
                # if type(cell) == "<class 'openpyxl.cell.cell.Cell'>":
                try:
                    # paste style to sheet receiving
                    if include_value:
                        cell.value = copiedData[countRow][countCol]

                    # print(type(cell))
                except:
                    pass

                cell.font = style[countRow][countCol][0]
                cell.border = style[countRow][countCol][1]
                cell.fill = style[countRow][countCol][2]
                cell.number_format = style[countRow][countCol][3]
                cell.protection = style[countRow][countCol][4]
                cell.alignment = style[countRow][countCol][5]

                countCol += 1

            countRow += 1

        return sheetReceiving

    def removeMerged(self, startCol, startRow, endCol, endRow, sheet):
        """usuwanie polaczen komorek z wybranego zakresu"""
        allMergedCells = sheet.merged_cells.ranges
        merged_list = []
        for ranges in allMergedCells:
            if ranges.min_col >= startCol and ranges.min_row >= startRow and ranges.max_col <= endCol and ranges.max_row <= endRow:
                merged_list.append([ranges.min_row, ranges.min_col, ranges.max_row, ranges.max_col])

        #druga petla jest zrobiona specjalnie i jest potrzebna poniewaz gdy komorki sa rozdzielane w pierwszej petli to losowo nie dziala
        for merged in merged_list:
            sheet.unmerge_cells(start_row=merged[0], start_column=merged[1], end_row=merged[2], end_column=merged[3])

        return sheet
    def run(self, dict_report, file_path, out_path, conv_errors_list):
        main_dir = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
        template_path = os.path.join(main_dir, 'raport', 'szablon_raport_z_importu.xlsx')

        control_sum_sha256 = self.controlSum(file_path)

        start_paste_row = 9
        template_wb = load_workbook(filename=template_path)


        sheet = template_wb['QMapa_import']
        header_egib_data, header_egib_style, header_egib_merged = self.copyRange(1,9,8,12,sheet)
        body_egib_data, body_egib_style, body_egib_merged = self.copyRange(1,13,8,13,sheet)
        bottom_egib_data, bottom_egib_style, bottom_egib_merged = self.copyRange(1,14,9,14,sheet)
        prez_egib_data, prez_egib_style, prez_egib_merged = self.copyRange(1,15,9,16,sheet)

        header_gesut_data, header_gesut_style, header_gesut_merged = self.copyRange(1,18,8,21,sheet)
        body_gesut_data, body_gesut_style, body_gesut_merged = self.copyRange(1,22,8,22,sheet)
        bottom_gesut_data, bottom_gesut_style, bottom_gesut_merged = self.copyRange(1,23,8,23,sheet)
        prez_gesut_data, prez_gesut_style, prez_gesut_merged = self.copyRange(1,24,9,25,sheet)


        header_bdot500_data, header_bdot500_style, header_bdot500_merged = self.copyRange(1,27,8,30,sheet)
        body_bdot500_data, body_bdot500_style, body_bdot500_merged = self.copyRange(1,31,8,31,sheet)
        bottom_bdot500_data, bottom_bdot500_style, bottom_bdot500_merged = self.copyRange(1,32,8,32,sheet)
        prez_bdot500_data, prez_bdot500_style, prez_bdot500_merged = self.copyRange(1,33,9,34,sheet)


        header_other_data, header_other_style, header_other_merged = self.copyRange(1,36,3,38,sheet)
        body_other_data, body_other_style, body_other_merged = self.copyRange(1,39,3,39,sheet)
        bottom_other_data, bottom_other_style, bottom_other_merged = self.copyRange(1,40,3,40,sheet)

        dict_copied_frame = {'EGiB':{'header':[header_egib_data,header_egib_style,header_egib_merged],'body':[body_egib_data,body_egib_style,body_egib_merged],'bottom':[bottom_egib_data,bottom_egib_style,bottom_egib_merged], 'prez':[prez_egib_data,prez_egib_style,prez_egib_merged]},'GESUT':{'header':[header_gesut_data,header_gesut_style,header_gesut_merged],'body':[body_gesut_data,body_gesut_style,body_gesut_merged],'bottom':[bottom_gesut_data,bottom_gesut_style,bottom_gesut_merged], 'prez': [prez_egib_data, prez_gesut_style, prez_gesut_merged]},'BDOT500':{'header':[header_bdot500_data,header_bdot500_style,header_bdot500_merged],'body':[body_bdot500_data,body_bdot500_style,body_bdot500_merged],'bottom':[bottom_bdot500_data,bottom_bdot500_style,bottom_bdot500_merged], 'prez': [prez_bdot500_data, prez_bdot500_style, prez_bdot500_merged]},'other':{'header':[header_other_data,header_other_style,header_other_merged],'body':[body_other_data,body_other_style,body_other_merged],'bottom':[bottom_other_data,bottom_other_style,bottom_other_merged]}}

        sheet = self.removeMerged(1, 9, 9, 41, sheet)
        sheet.delete_rows(9,41)

        '''sheet['A3'] = sheet['A3'].value + str(os.path.split(file_path)[1])
        sheet['A4'] = sheet['A4'].value + str(control_sum_sha256.hexdigest())
        sheet['A5'] = sheet['A5'].value + datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')'''

        sheet['C5'] = str(os.path.split(file_path)[1])
        sheet['C6'] = str(control_sum_sha256.hexdigest())
        sheet['C7'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        prez_layers = ['EGB_PrezentacjaGraficzna', 'OT_PrezentacjaGraficzna', 'GES_PrezentacjaGraficzna']


        for group_name, layers_info in dict_report.items():
            if len(layers_info) > 0:
                if group_name in ['EGiB', 'GESUT', 'BDOT500'] :
                    key_frame = group_name
                    last_column = 8
                    rows = 4  # header
                else:
                    key_frame = 'other'
                    last_column = 3
                    rows = 3  # header

                header_data = dict_copied_frame[key_frame]['header'][0]
                header_style = dict_copied_frame[key_frame]['header'][1]
                header_merged = dict_copied_frame[key_frame]['header'][2]

                body_data = dict_copied_frame[key_frame]['body'][0]
                body_style = dict_copied_frame[key_frame]['body'][1]
                body_merged = dict_copied_frame[key_frame]['body'][2]

                bottom_data = dict_copied_frame[key_frame]['bottom'][0]
                bottom_style = dict_copied_frame[key_frame]['bottom'][1]
                bottom_merged = dict_copied_frame[key_frame]['bottom'][2]

                #rows = 4
                sheet = self.pasteRange(1, start_paste_row, last_column, start_paste_row + rows - 1, sheet,
                                        header_data, header_style, header_merged)

                if key_frame == "other":
                    sheet.cell(column = 1, row = start_paste_row).value = str(sheet.cell(column = 1, row = start_paste_row).value) + ': '.join(group_name.split(': ')[1:])

                start_paste_row += rows

                i = 0
                sums = [0,0,0,0,0,0]
                short_sum = 0
                prez_info = None
                for layer_name, layer_inf in layers_info.items():
                    if layer_name in prez_layers and key_frame != "other":
                        prez_info = layer_inf
                        prez_name = layer_name
                        prez_data = dict_copied_frame[key_frame]['prez'][0]
                        prez_style = dict_copied_frame[key_frame]['prez'][1]
                        prez_merged = dict_copied_frame[key_frame]['prez'][2]
                    else:
                        i += 1
                        rows = 1
                        sheet = self.pasteRange(1, start_paste_row, last_column, start_paste_row+rows -1, sheet, body_data, body_style, body_merged)
                        sheet.cell(column=1, row=start_paste_row).value = i

                        if key_frame == "other" and incompatible_pref in layer_name:
                            # zmiana wstawianej warstwy poprzez usuniecie przedrostka z informacja o niezgodnosci i nazawa bazy
                            name_of_layer = '_'.join(layer_name.split('_')[3:])
                        else:
                            name_of_layer = layer_name
                        sheet.cell(column=2, row=start_paste_row).value = name_of_layer

                        if key_frame != "other":
                            sheet.cell(column=3, row=start_paste_row).value = sum(layer_inf[0:4])
                            sheet.cell(column=4, row=start_paste_row).value = layer_inf[0]
                            sheet.cell(column=5, row=start_paste_row).value = layer_inf[1]
                            sheet.cell(column=6, row=start_paste_row).value = layer_inf[2]
                            sheet.cell(column=7, row=start_paste_row).value = layer_inf[3]
                            sheet.cell(column=8, row=start_paste_row).value = layer_inf[4]

                            if sheet.cell(column=3, row=start_paste_row).value != sheet.cell(column=8, row=start_paste_row).value:
                                n_font = copy(sheet.cell(column=8, row=start_paste_row).font)
                                n_font.color = 'FFFF0000'
                                sheet.cell(column=8, row=start_paste_row).font = n_font


                            sums[0] += sum(layer_inf[0:4])
                            for nr in range(5):
                                sums[nr+1] += layer_inf[nr]

                        elif key_frame == "other":
                            sheet.cell(column=3, row=start_paste_row).value = layer_inf[0]
                            short_sum += layer_inf[0]


                        start_paste_row += rows


                rows = 1
                sheet = self.pasteRange(1, start_paste_row, last_column, start_paste_row + rows - 1, sheet,
                                        bottom_data, bottom_style, bottom_merged)

                if key_frame != "other":
                    for nr in range(6):
                        sheet.cell(column=3+nr, row=start_paste_row).value = sums[nr]
                elif key_frame == "other":
                    sheet.cell(column=3, row=start_paste_row).value = short_sum

                start_paste_row += rows

                if prez_info is not None:
                    rows = 2
                    sheet = self.pasteRange(1, start_paste_row, last_column, start_paste_row + rows - 1, sheet,
                                            prez_data, prez_style, prez_merged)
                    sheet.cell(column = 2, row = start_paste_row + 1).value = prez_name
                    sheet.cell(column = 8, row=start_paste_row + 1).value = prez_info[0]
                    start_paste_row += rows

                start_paste_row += 1  # zrobienie odstepu 1 wiersza

        # Dodanie info o bledach importu
        sheet_errors = template_wb['Błędy_importu']
        if len(conv_errors_list) > 0:
            sheet.cell(column=1, row=start_paste_row).value = 'W czasie importu wystąpiły błędy. Niektóre obiekty mogły nie zostać zaimportowane. Szczegóły na następnym arkuszu.'
            sheet.cell(column=1, row=start_paste_row).alignment = Alignment(horizontal='left')

            n_font = copy(sheet.cell(column=1, row=start_paste_row).font)
            n_font.color = 'FFFF0000'
            sheet.cell(column=1, row=start_paste_row).font = n_font

            row_err = 1
            for conv_err in conv_errors_list:
                sheet_errors.cell(column=1, row=row_err).value = conv_err
                row_err += 1
        else:
            template_wb.remove(sheet_errors)

        #file_report_path = file_path + '_raport.xlsx'
        file_report_path = out_path

        template_wb.save(file_report_path)

        return(file_report_path)
